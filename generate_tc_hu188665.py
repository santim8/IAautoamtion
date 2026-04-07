"""
Script puntual: genera el Excel de casos de prueba para la HU 188665
Aumentos y Reactivaciones de Cupo - Feature Flujo Autogestionado
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path

OUTPUT_DIR = Path(r"C:\Users\santiago.correa03\ColsubsidioIA\IAautoamtion\excels")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

HEADER_BG   = "1F4E79"
HEADER_FONT = "FFFFFF"
TC_ROW_BG   = "D6E4F0"
AREA_PATH   = "Ecosistema Digital Crédito y Seguros"
STATE       = "Design"
HEADERS     = ["ID","Work Item Type","Title","Test Step",
               "Step Action","Step Expected","Area Path","Assigned To","State"]
COL_WIDTHS  = {"A":10,"B":16,"C":70,"D":11,"E":70,"F":60,"G":38,"H":13,"I":10}

test_cases = [
    {
        "title": "TC001 - Visualización de pantalla de oferta para Reactivación Simple aprobada",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado en el flujo Autogestionado\n"
                    "- Motor de decisión retorna novedad tipo 'Reactivación Simple' con estado aprobado\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor de decisión retorna una oferta aprobada de tipo Reactivación Simple", "expected": ""},
            {"action": "When el usuario accede a la pantalla de configuración de oferta", "expected": ""},
            {"action": "Then el título de la pantalla indica que la solicitud de reactivación fue aprobada",
             "expected": "El título refleja explícitamente la aprobación de la reactivación"},
            {"action": "And el cupo actual del usuario se muestra visible en pantalla",
             "expected": "Se muestra el mismo cupo vigente sin cambio de valor"},
            {"action": "And el componente 'Ajustar cupo' está deshabilitado o ausente",
             "expected": "No es posible modificar el monto del cupo"},
            {"action": "And la pantalla de selección de quincena está oculta o bloqueada",
             "expected": "No se puede interactuar con la selección de quincena"},
            {"action": "And se muestran las condiciones vigentes: fecha de pago, cuota de manejo, seguros y sistema de financiación",
             "expected": "Toda la información de condiciones es visible y correcta"},
            {"action": "And se conserva la pregunta '¿Qué pasa si no uso mi Cupo de Crédito?'",
             "expected": "Sección visible en pantalla"},
            {"action": "And solo están habilitadas las acciones 'Aceptar el cupo de crédito' y 'Abandonar la solicitud'",
             "expected": "Las demás acciones están inactivas o no visibles"},
        ]
    },
    {
        "title": "TC002 - Ajuste de cupo dentro del rango permitido para Aumento puro",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado con cupo actual vigente (ej. $1.000.000)\n"
                    "- Motor retorna novedad tipo 'Aumento' con oferta máxima aprobada (ej. $2.000.000)\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor de decisión retorna una oferta aprobada de tipo Aumento", "expected": ""},
            {"action": "When el usuario presiona la opción 'Ajustar cupo'",
             "expected": "El componente de ajuste de cupo se activa y es editable"},
            {"action": "Then el sistema muestra el valor mínimo permitido igual al cupo actual del usuario",
             "expected": "El valor mínimo corresponde al cupo actual vigente del afiliado"},
            {"action": "And el sistema muestra el valor máximo permitido igual al monto aprobado por el motor",
             "expected": "El valor máximo corresponde a la oferta aprobada por el motor"},
            {"action": "When el usuario ingresa un valor dentro del rango permitido (ej. $1.500.000)", "expected": ""},
            {"action": "Then el sistema acepta el valor sin mostrar error de validación",
             "expected": "No se muestra mensaje de error; el valor es aceptado"},
            {"action": "And si el sistema de financiación es Fijo la cuota fija se recalcula automáticamente",
             "expected": "La cuota fija refleja el nuevo monto ingresado de forma dinámica"},
            {"action": "And la pantalla de selección de quincena está oculta o bloqueada",
             "expected": "No se puede interactuar con la selección de quincena"},
        ]
    },
    {
        "title": "TC003 - Validación de límite inferior al ajustar cupo para Aumento (negativo)",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado con cupo actual (ej. $1.000.000)\n"
                    "- Motor retorna Aumento con oferta máxima (ej. $2.000.000)\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor retorna una oferta aprobada de tipo Aumento", "expected": ""},
            {"action": "When el usuario accede al componente de ajuste de cupo", "expected": ""},
            {"action": "And el usuario ingresa un valor inferior al cupo actual (ej. $800.000)", "expected": ""},
            {"action": "Then el sistema muestra un mensaje de validación indicando que el valor no puede ser menor al cupo actual",
             "expected": "Mensaje de error visible con indicación del valor mínimo permitido"},
            {"action": "And el sistema no permite continuar ni enviar el formulario con ese valor",
             "expected": "El botón de aceptar está deshabilitado o el flujo queda bloqueado"},
        ]
    },
    {
        "title": "TC004 - Validación de límite superior al ajustar cupo para Aumento (negativo)",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado con cupo actual (ej. $1.000.000)\n"
                    "- Motor retorna Aumento con oferta máxima (ej. $2.000.000)\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor retorna una oferta aprobada de tipo Aumento", "expected": ""},
            {"action": "When el usuario accede al componente de ajuste de cupo", "expected": ""},
            {"action": "And el usuario ingresa un valor superior al monto aprobado por el motor (ej. $2.500.000)", "expected": ""},
            {"action": "Then el sistema muestra un mensaje de validación indicando que el valor no puede superar la oferta aprobada",
             "expected": "Mensaje de error visible con indicación del valor máximo permitido"},
            {"action": "And el sistema no permite continuar ni enviar el formulario con ese valor",
             "expected": "El botón de aceptar está deshabilitado o el flujo queda bloqueado"},
        ]
    },
    {
        "title": "TC005 - Configuración de pantalla para Reactivación con Aumento simultáneo",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado\n"
                    "- Motor retorna novedad 'Reactivación con Aumento' vía Redeban\n"
                    "- La reactivación y el aumento no se aplican en el mismo día (regla Redeban)\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor retorna una oferta aprobada que incluye Reactivación y Aumento simultáneo", "expected": ""},
            {"action": "When el usuario accede a la pantalla de configuración de oferta", "expected": ""},
            {"action": "Then el título indica explícitamente la aprobación de la reactivación y ofrece el beneficio adicional del aumento",
             "expected": "Título diferenciado respecto a los escenarios de Reactivación Simple y Aumento puro"},
            {"action": "And se muestra el cupo aprobado incluyendo el valor del aumento",
             "expected": "El valor mostrado corresponde al aprobado por el motor incluyendo el aumento"},
            {"action": "And el componente de ajuste de cupo está habilitado",
             "expected": "El usuario puede interactuar con el componente de ajuste"},
            {"action": "And el límite mínimo del ajuste es el cupo con el que el usuario realizó la solicitud",
             "expected": "El sistema muestra y valida correctamente el valor mínimo"},
            {"action": "And el límite máximo es el cupo ofertado por el motor",
             "expected": "El sistema muestra y valida correctamente el valor máximo"},
            {"action": "And la pantalla de selección de quincena está oculta o bloqueada",
             "expected": "No se puede modificar la quincena"},
            {"action": "And se conserva la alerta informativa sin modificaciones",
             "expected": "La alerta se muestra tal como está actualmente definida"},
            {"action": "And se conserva la pregunta '¿Qué pasa si no uso mi Cupo de Crédito?'",
             "expected": "Sección visible en pantalla"},
        ]
    },
    {
        "title": "TC006 - Pantalla final de éxito para Aumento aceptado",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado\n"
                    "- Ha aceptado exitosamente una oferta de tipo Aumento\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el usuario acepta exitosamente una oferta de tipo Aumento", "expected": ""},
            {"action": "When el sistema redirige al usuario a la pantalla final de éxito", "expected": ""},
            {"action": "Then se muestra el mensaje informando que recibirá los documentos legales referentes al aumento",
             "expected": "Mensaje sobre documentos legales visible en pantalla"},
            {"action": "And el mensaje orienta al usuario a utilizar su cupo en los convenios disponibles",
             "expected": "Texto de orientación de uso en convenios visible"},
            {"action": "And no se muestra el texto estándar de cupo de crédito correspondiente a Reactivación",
             "expected": "El contenido es diferenciado y exclusivo para el escenario de Aumento"},
        ]
    },
    {
        "title": "TC007 - Pantalla final de éxito para Reactivación Simple aceptada",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado\n"
                    "- Ha aceptado exitosamente una oferta de tipo Reactivación Simple\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el usuario acepta exitosamente una oferta de tipo Reactivación Simple", "expected": ""},
            {"action": "When el sistema redirige al usuario a la pantalla final de éxito", "expected": ""},
            {"action": "Then se muestra el texto estándar de cupo de crédito sin modificaciones",
             "expected": "Contenido estándar visible e idéntico al flujo normal de cupo"},
            {"action": "And no se muestra el mensaje de documentos legales propio del Aumento",
             "expected": "Contenido específico de Aumento no visible en pantalla"},
        ]
    },
    {
        "title": "TC008 - Pantalla final con próximos pasos para Reactivación con Aumento",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado\n"
                    "- Ha aceptado una oferta de tipo Reactivación con Aumento\n"
                    "- Regla Redeban: la reactivación y el aumento no se aplican el mismo día\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el usuario acepta exitosamente una oferta de tipo Reactivación con Aumento", "expected": ""},
            {"action": "When el sistema redirige al usuario a la pantalla final", "expected": ""},
            {"action": "Then se informa al usuario sobre los próximos pasos relacionados con el aumento",
             "expected": "Sección de próximos pasos visible en pantalla"},
            {"action": "And el mensaje es claro respecto a que el aumento no se aplica el mismo día de la reactivación",
             "expected": "El usuario comprende que el aumento se aplica en un momento posterior a la reactivación"},
        ]
    },
    {
        "title": "TC009 - Negación de Reactivación Simple por el motor de decisión",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado con solicitud activa de tipo Reactivación Simple\n"
                    "- Motor de decisión emite respuesta de negación\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor de decisión niega la solicitud de Reactivación Simple", "expected": ""},
            {"action": "When el sistema procesa la respuesta de negación del motor", "expected": ""},
            {"action": "Then se muestra un mensaje en pantalla informando la negación al usuario",
             "expected": "Mensaje de negación claro y comprensible para el usuario"},
            {"action": "And el sistema impide al usuario continuar con la solicitud de reactivación",
             "expected": "El flujo queda bloqueado tras la negación"},
            {"action": "And no se muestran opciones para continuar el proceso de reactivación",
             "expected": "Botones de continuación deshabilitados o no visibles"},
        ]
    },
    {
        "title": "TC010 - Negación de Aumento con invitación al uso del cupo actual",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado con solicitud activa de tipo Aumento\n"
                    "- Motor de decisión emite respuesta de negación\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor de decisión niega la solicitud de Aumento de cupo", "expected": ""},
            {"action": "When el sistema procesa la respuesta de negación del motor", "expected": ""},
            {"action": "Then se muestra un mensaje en pantalla informando la decisión de negación",
             "expected": "Mensaje de negación visible al usuario"},
            {"action": "And se incluye un texto que invita al usuario a continuar usando su cupo actual",
             "expected": "Texto de invitación al uso del cupo vigente visible en pantalla"},
            {"action": "And el sistema impide al usuario continuar con la solicitud de aumento",
             "expected": "El flujo queda bloqueado; no hay opción de continuar con el aumento"},
        ]
    },
    {
        "title": "TC011 - Validación de integridad en el envío de datos de aceptación de Aumento",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado\n"
                    "- Ha ajustado el cupo a un valor dentro del rango permitido\n"
                    "- Listo para confirmar la aceptación\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el usuario ha seleccionado un monto de cupo válido para un Aumento", "expected": ""},
            {"action": "When el usuario confirma la aceptación del cupo", "expected": ""},
            {"action": "Then el sistema valida en backend que el valor enviado no sea menor al cupo actual",
             "expected": "Validación de integridad ejecutada correctamente"},
            {"action": "And el sistema valida en backend que el valor enviado no sea mayor al cupo otorgado por el motor",
             "expected": "El valor respeta el tope aprobado"},
            {"action": "And si los valores son válidos el sistema procesa la aceptación exitosamente",
             "expected": "Solicitud procesada y flujo avanza a pantalla de éxito"},
        ]
    },
    {
        "title": "TC012 - Recálculo dinámico de cuota fija en sistema Fijo para Reactivación con Aumento",
        "steps": [
            {
                "action": (
                    "Precondiciones:\n"
                    "- Usuario autenticado con sistema de financiación tipo 'Fijo'\n"
                    "- Motor retorna Reactivación con Aumento\n"
                    "- Ambiente: QA/DEV"
                ),
                "expected": ""
            },
            {"action": "Given el motor retorna una oferta de Reactivación con Aumento y el usuario tiene sistema de financiación Fijo",
             "expected": ""},
            {"action": "When el usuario ajusta el monto del cupo dentro del rango permitido", "expected": ""},
            {"action": "Then la cuota fija se recalcula automáticamente y muestra el nuevo valor",
             "expected": "El valor de cuota fija se actualiza en tiempo real al cambiar el monto"},
            {"action": "And el cambio en la cuota fija es visible y comprensible para el afiliado antes de aceptar",
             "expected": "El afiliado puede identificar claramente el impacto del aumento en su cuota fija"},
            {"action": "And al aceptar el sistema envía el valor de cuota fija consistente con el cupo seleccionado",
             "expected": "Los datos enviados son coherentes entre monto de cupo y cuota fija recalculada"},
        ]
    },
]


def _header_style(cell):
    cell.font      = Font(bold=True, color=HEADER_FONT)
    cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _tc_row_style(cell, bold=False):
    cell.fill      = PatternFill("solid", fgColor=TC_ROW_BG)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font      = Font(size=11, bold=bold)


def _step_row_style(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font      = Font(size=11)


def write_excel(cases, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    for col_idx, header in enumerate(HEADERS, start=1):
        _header_style(ws.cell(row=1, column=col_idx, value=header))

    current_row = 2
    for tc in cases:
        # TC header row
        for col_letter, value, bold in [
            ("B", "Test Case", False),
            ("C", tc["title"], True),
            ("G", AREA_PATH, False),
            ("I", STATE, False),
        ]:
            col_idx = ord(col_letter) - ord("A") + 1
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            _tc_row_style(cell, bold=bold)
        current_row += 1

        # Step rows
        for step_num, step in enumerate(tc["steps"], start=1):
            for col_letter, value in [
                ("D", float(step_num)),
                ("E", step["action"]),
                ("F", step["expected"] or None),
            ]:
                col_idx = ord(col_letter) - ord("A") + 1
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                _step_row_style(cell)
            current_row += 1

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(out_path)


if __name__ == "__main__":
    stamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = OUTPUT_DIR / f"TC_Aumentos_Reactivaciones_HU188665_{stamp}.xlsx"
    write_excel(test_cases, out_file)
    total_steps = sum(len(tc["steps"]) for tc in test_cases)
    print(f"\n Excel generado: {out_file}")
    print(f"  Casos de prueba : {len(test_cases)}")
    print(f"  Total de pasos  : {total_steps}")
