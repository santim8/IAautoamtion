"""
create_testcases_automation.py
-------------------------------
Genera un Excel de casos de prueba a partir de una Historia de Usuario.
El Excel replica exactamente el template de Colsubsidio.

ARCHIVOS FIJOS (recomendado)
----------------------------
  Entrada:  test_cases.json  (en la raíz del repo — se reemplaza por cada HU)
  Salida:   test_cases.xlsx  (en la raíz del repo — se sobreescribe)

  Modo por defecto (sin argumentos): lee test_cases.json y genera test_cases.xlsx
    python create_testcases_automation.py

MODOS DE USO
------------
  Modo 1 – HU en texto/Excel → OpenAI genera los TCs:
    python create_testcases_automation.py historia.txt

  Modo 2 – JSON personalizado:
    python create_testcases_automation.py mi_archivo.json

  Carpeta de salida personalizada (opcional):
    python create_testcases_automation.py --output "C:\\MiCarpeta"

VARIABLES DE ENTORNO
--------------------
  OPENAI_API_KEY    → clave de OpenAI         (requerida solo en Modo 1)
  OPENAI_MODEL      → modelo a usar           (por defecto: gpt-4.1-mini)
  BIZAGI_AREA_PATH  → área en Azure DevOps    (por defecto: Ecosistema Digital Crédito y Seguros)
  BIZAGI_ASSIGNED_TO→ responsable del TC      (por defecto: vacío)
  TC_OUTPUT_DIR     → carpeta de salida       (por defecto: raíz del repo)

COLUMNAS DEL EXCEL
------------------
  A=ID | B=Work Item Type | C=Title | D=Test Step |
  E=Step Action | F=Step Expected | G=Area Path | H=Assigned To | I=State
"""

import os
import sys
import json
import logging
import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ---------------------------
# CONFIG
# ---------------------------
# Script directory — works no matter from where it is called
SCRIPT_DIR = Path(__file__).resolve().parent

# Output folder: env var > default root of repo (same folder as the script)
_env_output = os.getenv("TC_OUTPUT_DIR")
OUTPUT_DIR  = Path(_env_output) if _env_output else SCRIPT_DIR
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Fixed file names (replace on every story)
DEFAULT_INPUT_JSON = SCRIPT_DIR / "test_cases.json"
DEFAULT_OUTPUT_XLSX = "test_cases.xlsx"

OPENAI_MODEL  = os.getenv("OPENAI_MODEL", "gpt-4.1-mini")
AREA_PATH     = os.getenv("BIZAGI_AREA_PATH", "Ecosistema Digital Crédito y Seguros")
ASSIGNED_TO   = os.getenv("BIZAGI_ASSIGNED_TO", "")
DEFAULT_STATE = "Design"

# Excel header colors (matching template)
HEADER_BG    = "1F4E79"   # dark blue
HEADER_FONT  = "FFFFFF"   # white
TC_ROW_BG    = "D6E4F0"   # light blue for test case rows

# Column widths matching the reference file
COLUMN_WIDTHS = {
    "A": 10,   # ID
    "B": 16,   # Work Item Type
    "C": 60,   # Title
    "D": 11,   # Test Step
    "E": 60,   # Step Action
    "F": 55,   # Step Expected
    "G": 38,   # Area Path
    "H": 13,   # Assigned To
    "I": 10,   # State
}

HEADERS = ["ID", "Work Item Type", "Title", "Test Step",
           "Step Action", "Step Expected", "Area Path", "Assigned To", "State"]

# ---------------------------
# INPUT READERS
# ---------------------------
def read_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def read_xlsx_as_text(path: str, max_rows_per_sheet: int = 200) -> str:
    try:
        xls = pd.ExcelFile(path)
        chunks = [f"### FILE: {Path(path).name}"]
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, dtype=str).fillna("")
            chunks.append(f"\n# SHEET: {sheet}\n{df.head(max_rows_per_sheet).to_csv(index=False)}")
        return "\n".join(chunks)
    except Exception as e:
        return f"### FILE: {Path(path).name}\n[Error al leer: {e}]"

def load_user_story(path: str) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Archivo no encontrado: {path}")
    if path.lower().endswith(".xlsx"):
        return read_xlsx_as_text(path)
    return read_txt(path)

def load_json_testcases(path: str) -> list[dict[str, Any]]:
    """Carga casos de prueba desde un JSON previamente generado por cualquier IA."""
    with open(path, "r", encoding="utf-8") as f:
        raw = f.read().strip()
    # Limpia bloques markdown si la IA los incluyó
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()
    data = json.loads(raw)
    if not isinstance(data, list):
        raise RuntimeError("El JSON debe ser un array de casos de prueba.")
    return data

# ---------------------------
# AI: GENERATE TEST CASES
# ---------------------------
def build_prompt(user_story: str) -> str:
    return f"""
Eres un QA Senior experto en pruebas funcionales para aplicaciones financieras.
A partir de la Historia de Usuario que se te proporciona, genera casos de prueba detallados.

INSTRUCCIONES:
- Genera casos positivos, negativos y de borde según las reglas de negocio.
- El primer paso de CADA caso SIEMPRE debe ser las "Precondiciones".
- Los pasos siguientes usan Given / When / Then / And / But.
- El campo "expected" es el resultado esperado; puede quedar vacío en pasos intermedios.
- Devuelve ÚNICAMENTE JSON válido, sin texto adicional, sin bloques de código markdown.

FORMATO DE SALIDA (array JSON):
[
  {{
    "title": "TC001 - Descripción corta del caso",
    "steps": [
      {{"action": "Precondiciones:\\n- condición 1\\n- condición 2", "expected": ""}},
      {{"action": "Given <contexto inicial>", "expected": ""}},
      {{"action": "When <acción del usuario>", "expected": ""}},
      {{"action": "Then <validación principal>", "expected": "<resultado esperado>"}},
      {{"action": "And <validación adicional>", "expected": "<resultado esperado>"}}
    ]
  }}
]

HISTORIA DE USUARIO:
{user_story}
""".strip()

def call_openai(prompt: str) -> list[dict[str, Any]]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Falta OPENAI_API_KEY en variables de entorno.")

    from openai import OpenAI
    client = OpenAI(api_key=api_key)

    log.info("Llamando a OpenAI (%s)…", OPENAI_MODEL)
    resp = client.responses.create(
        model=OPENAI_MODEL,
        input=prompt,
        temperature=0.2,
    )
    raw = resp.output_text.strip()

    # Strip markdown code fences if the model adds them
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("La IA no devolvió JSON válido:\n%s", raw[:500])
        raise RuntimeError(f"JSON inválido de la IA: {e}") from e

    if not isinstance(data, list):
        raise RuntimeError("Se esperaba un array JSON en la respuesta de la IA.")

    return data

# ---------------------------
# EXCEL WRITER
# ---------------------------
def _header_style(cell) -> None:
    cell.font      = Font(bold=True, color=HEADER_FONT)
    cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

def _tc_row_style(cell) -> None:
    cell.fill      = PatternFill("solid", fgColor=TC_ROW_BG)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font      = Font(size=11)

def _step_row_style(cell) -> None:
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font      = Font(size=11)

def write_excel(test_cases: list[dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _header_style(cell)

    # ── Data rows ────────────────────────────────────────────────────────────
    current_row = 2
    for tc_idx, tc in enumerate(test_cases, start=1):
        title  = tc.get("title", f"TC{tc_idx:03d} - Sin título")
        steps  = tc.get("steps", [])

        # Ensure TC number prefix is present
        if not title.upper().startswith("TC"):
            title = f"TC{tc_idx:03d} - {title}"

        # Test Case header row
        tc_data = {
            "B": "Test Case",
            "C": title,
            "G": AREA_PATH,
            "H": ASSIGNED_TO,
            "I": DEFAULT_STATE,
        }
        for col_letter, value in tc_data.items():
            col_idx = ord(col_letter) - ord("A") + 1
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            _tc_row_style(cell)
        current_row += 1

        # Step rows
        for step_num, step in enumerate(steps, start=1):
            action   = step.get("action", "").strip()
            expected = step.get("expected", "").strip()

            step_data = {
                "D": float(step_num),
                "E": action,
                "F": expected if expected else None,
            }
            for col_letter, value in step_data.items():
                col_idx = ord(col_letter) - ord("A") + 1
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                _step_row_style(cell)
            current_row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(out_path)
    log.info("Excel guardado en: %s", out_path)

# ---------------------------
# MAIN
# ---------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Genera Excel de casos de prueba desde una HU o un JSON de IA.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Ejemplos:\n"
            "  # Por defecto: lee test_cases.json y genera test_cases.xlsx:\n"
            '  python create_testcases_automation.py\n\n'
            "  # Modo 1 — HU en .txt, OpenAI genera los TCs:\n"
            '  python create_testcases_automation.py historia.txt\n\n'
            "  # Modo 2 — JSON personalizado:\n"
            '  python create_testcases_automation.py mi_archivo.json\n\n'
            "  # Con carpeta de salida personalizada:\n"
            '  python create_testcases_automation.py --output "C:\\MiCarpeta"\n'
        ),
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=str(DEFAULT_INPUT_JSON),
        help="Ruta al archivo de entrada (.txt, .xlsx, .json). Por defecto: test_cases.json",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Carpeta donde guardar el Excel (por defecto: raíz del repo).",
    )
    args = parser.parse_args()

    # Resolve output directory
    out_dir = Path(args.output).resolve() if args.output else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    input_path = args.input

    # ── Decide mode based on extension ──────────────────────────────────────
    if input_path.lower().endswith(".json"):
        log.info("Modo JSON: cargando casos de prueba desde %s", input_path)
        test_cases = load_json_testcases(input_path)
    else:
        log.info("Modo HU: leyendo historia de usuario desde %s", input_path)
        user_story = load_user_story(input_path)
        prompt     = build_prompt(user_story)
        test_cases = call_openai(prompt)

    if not test_cases:
        raise RuntimeError("No se encontraron casos de prueba en la entrada.")

    log.info("Casos de prueba a escribir: %d", len(test_cases))

    out_file = out_dir / DEFAULT_OUTPUT_XLSX
    write_excel(test_cases, out_file)

    total_steps = sum(len(tc.get("steps", [])) for tc in test_cases)
    print(f"\n  Excel generado : {out_file}")
    print(f"  Casos de prueba: {len(test_cases)}")
    print(f"  Total de pasos : {total_steps}")


if __name__ == "__main__":
    main()
