#!/usr/bin/env python3
"""Panel local para lanzar los scripts del repo sin tocar la consola.

Cada herramienta es una entrada de HERRAMIENTAS: se declara el script, los
campos que pide y como se para. El panel arma el formulario, el comando y el
log solo. Agregar una herramienta nueva es agregar un diccionario, no tocar la
UI.

Los tres modos de parada, que es lo que de verdad separa a un script de otro:

  "centinela"  el proceso escribe evidencia al cerrarse (observador_flujo).
               Se le pide parar con un archivo que el vigila, para que salga
               por el mismo camino que Ctrl+C. Matarlo dejaria la corrida sin
               reporte, sin el ultimo pantallazo y sin validacion de esquemas.
  "terminar"   el proceso queda vivo sosteniendo algo (el navegador de
               bizagi_consultar_caso, que espera un Enter que nunca llega
               cuando no hay consola). No hay nada que perder: se termina.
  None         de un tiro: corre, termina y se le lee un veredicto del log.

Se abre con doble clic en panel.bat, o con:
    python panel_observador.py
"""
import json
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
import time
import tkinter as tk
import urllib.error
import urllib.request
import webbrowser
from tkinter import ttk, messagebox

AQUI = os.path.dirname(os.path.abspath(__file__))
EVIDENCIAS = os.path.join(AQUI, "evidences")


def _config():
    """Ajustes por maquina, en panel.config.json (no se versiona)."""
    try:
        with open(os.path.join(AQUI, "panel.config.json"), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}


CONFIG = _config()
# El framework Java vive en otro repo y cada quien lo clona donde quiere: se
# puede fijar por entorno o por panel.config.json antes de caer al default.
FRAMEWORK = (os.environ.get("COLSUBSIDIO_FRAMEWORK")
             or CONFIG.get("framework")
             or os.path.join(os.path.expanduser("~"), "IdeaProjects",
                             "colsubsidioFramework"))
SUITE_BIOMETRIA = os.path.join("src", "test", "resources", "suits", "testng-biometry.xml")
# Usuarios de prueba. Viven FUERA del repo a proposito: dentro, un git clean
# -fdx o volver a clonar se los lleva por delante (estan en .gitignore, que es
# justo lo que git clean borra). Aqui sobreviven a cualquier cosa que le pase
# al repo.
DIR_DATOS = os.path.join(os.path.expanduser("~"), ".panel_qa")
USUARIOS = os.path.join(DIR_DATOS, "usuarios_prueba.json")
USUARIOS_LEGADO = os.path.join(AQUI, "usuarios_prueba.json")
# Catalogo compartido: este SI se versiona, para que quien clone el repo
# arranque con los usuarios de prueba del equipo. Siembra el archivo local la
# primera vez; despues cada quien maneja el suyo y sincroniza a mano.
USUARIOS_COMPARTIDOS = os.path.join(AQUI, "usuarios_compartidos.json")
BACKUPS = os.path.join(DIR_DATOS, "backups")
LOGS = os.path.join(DIR_DATOS, "logs")
MAX_BACKUPS = 30

DATA_PROVIDER = os.path.join("src", "test", "java", "execution", "data",
                             "DataProviderUtil.java")
# Object[][] raw = { ... }; del DataProvider fillDataApi. El bloque de
# deduplicacion que sigue (LinkedHashSet) no se toca: dedupe en runtime.
RE_RAW = re.compile(r"        Object\[\]\[\] raw = \{.*?\n        \};", re.S)
# [EXCEL_UPDATE] Thread:28 52526685 - Card Validation - PASS - HTTP 200 estado OK
RE_EXCEL = re.compile(r"\[EXCEL_UPDATE\] Thread:\d+\s+(\S+)\s+-\s+(.+?)\s+-\s+(.+)")
# Servicios del catalogo del framework que no se revisan aqui: se descartan al
# entrar, asi no ocupan columna ni en la tabla ni en el Excel. En minusculas
# porque el log no siempre respeta el casing del catalogo.
SERVICIOS_OCULTOS = {"sso credentials", "habeas data records", "validate request"}
# Orden canonico de las columnas, copiado de SERVICE_COLUMNS en
# ExcelReportManager.java, para que la tabla del panel se lea igual que el Excel
# del framework en vez de por orden de llegada al log.
SERVICIOS_CATALOGO = [
    "SSO Credentials", "Validation Bizagi", "Validator Rights", "Card Number",
    "Card Status", "Salary", "Card Validation", "Card Validation Error",
    "Novedad Estado", "Card Validation Error V2", "Restrictive List",
    "Siif Validation", "Validate Request", "Preapproved", "Habeas Data Records",
]
# Columnas que se muestran siempre, traigan o no linea en el log: Card
# Validation Error solo se reporta cuando falla, y que venga vacia tambien dice
# algo. Agrega aqui cualquier otra que quieras ver fija.
SERVICIOS_FIJOS = ["Card Validation Error"]


def _orden_servicio(nombre):
    """Coloca cada servicio donde lo pone el Excel; lo desconocido, al final."""
    try:
        return (SERVICIOS_CATALOGO.index(nombre.strip()), "")
    except ValueError:
        return (len(SERVICIOS_CATALOGO), nombre)
STOP_FILE = os.path.join(AQUI, ".detener_observador")
# Centinela propio: la analitica y el observador de red pueden correr a la
# vez sobre la misma pestana, y cada uno tiene que poder pararse solo.
STOP_FILE_ANALITICA = os.path.join(AQUI, ".detener_analitica")
# Lo que se escribe dentro del centinela para pedir parada sin reporte; el
# observador lo lee y deja la evidencia cruda para reportarla despues.
SIN_REPORTE = "sin-reporte"
PUERTO = 9222

RE_MARCA = re.compile(r"_(\d{4})-(\d{2})-(\d{2})_(\d{2})(\d{2})(\d{2})$")
RE_CASO = re.compile(r"ltima solicitud:\s*(\d+)")


# --- pre-flight y veredictos ----------------------------------------------
def chrome_arriba(puerto=PUERTO, timeout=1.0):
    """Chrome ya expone el puerto CDP? Es lo que decide si hay que lanzarlo."""
    try:
        with urllib.request.urlopen("http://localhost:%d/json/version" % puerto,
                                    timeout=timeout):
            return True
    except (urllib.error.URLError, OSError):
        return False


def preparar_chrome(emitir):
    """El observador se engancha a un Chrome ya abierto; si no hay, lo abre."""
    if chrome_arriba():
        emitir("> Chrome ya estaba arriba, lo reuso.", "panel")
        return True
    emitir("> Chrome no responde en el puerto %d, lo abro..." % PUERTO, "panel")
    subprocess.run([sys.executable, os.path.join(AQUI, "observador_flujo.py"),
                    "--lanzar-chrome"], cwd=AQUI, capture_output=True, text=True)
    for _ in range(40):          # hasta 20 s
        if chrome_arriba():
            emitir("> Chrome listo.", "panel")
            return True
        time.sleep(0.5)
    emitir("! Chrome no levanto el puerto de depuracion.", "mal")
    return False


# Los veredictos se buscan por subcadenas SIN tildes: el log del hijo puede
# llegar con la codificacion estropeada y "cancelación" no siempre casa.
def estado_cancelacion(texto):
    """bizagi_cancel_case no devuelve exit code distinto: el veredicto se lee."""
    m = RE_CASO.search(texto)
    caso = m.group(1) if m else "?"
    if "n aceptada para caso" in texto:
        return "bien", "Caso %s cancelado." % caso
    if "Error al cancelar el caso" in texto:
        return "mal", ("No se encontro el caso %s para cancelar "
                       "(ya cancelado o en un estado que no lo permite)." % caso)
    if "No se pudo obtener" in texto:
        return "mal", "No se encontro ninguna solicitud para ese documento."
    return "", "Termino sin un veredicto claro; revisa el log."


def estado_consulta(texto):
    m = RE_CASO.search(texto)
    if m:
        return "bien", "Ultima solicitud: %s" % m.group(1)
    return "mal", "No se encontro ninguna solicitud para ese documento."


def args_biometria(valores, emitir):
    """La suite de TestNG trae idCaso/typeDocument/identification fijos en el XML
    y el test los lee con @Parameters, que no se pueden pisar con -D.

    En vez de tocar el framework, se escribe una copia de la suite con los
    valores del formulario y se corre esa. Va a target/, que esta en .gitignore,
    para no ensuciar el repo del framework.
    """
    plantilla = os.path.join(FRAMEWORK, SUITE_BIOMETRIA)
    if not os.path.exists(plantilla):
        emitir("! No encuentro la suite: %s" % plantilla, "mal")
        return None
    with open(plantilla, encoding="utf-8") as f:
        xml = f.read()
    cambios = {"idCaso": valores["Id caso"], "typeDocument": valores["Tipo doc"],
               "identification": valores["Documento"]}
    for nombre, valor in cambios.items():
        xml, n = re.subn(r'(<parameter\s+name="%s"\s+value=")[^"]*(")' % nombre,
                         lambda m: m.group(1) + valor + m.group(2), xml)
        if not n:
            emitir("! La suite no declara el parametro %s" % nombre, "mal")
            return None
    destino_abs = os.path.join(FRAMEWORK, "target", "panel-suites")
    os.makedirs(destino_abs, exist_ok=True)
    ruta = os.path.join(destino_abs, "testng-biometry-panel.xml")
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(xml)
    emitir("> suite: caso %s, %s %s" % (cambios["idCaso"], cambios["typeDocument"],
                                        cambios["identification"]), "panel")
    return ["-DsuiteXmlFile=" + os.path.relpath(ruta, FRAMEWORK).replace("\\", "/")]


def mvn():
    """En Windows el ejecutable es mvn.cmd; el 'mvn' pelado es el script de sh."""
    return shutil.which("mvn.cmd") or shutil.which("mvn") or "mvn"


# --- registro de herramientas ---------------------------------------------
# tipo: "texto" | "opcion" | "check".  arg=None => argumento posicional.
HERRAMIENTAS = [
    {
        "id": "observador",
        "nombre": "Observador de flujos",
        "script": "observador_flujo.py",
        "boton": "Iniciar captura",
        "parada": "centinela",
        "stop_file": STOP_FILE,
        "previo": preparar_chrome,
        "ayuda": "Navega a mano en Chrome; captura pantallas, requests y websocket.",
        "campos": [
            {"tipo": "texto", "arg": "--flujo", "etiqueta": "Flujo",
             "valor": "login-credito", "ancho": 26},
            {"tipo": "texto", "arg": "--solo-url", "etiqueta": "Solo URL que contenga",
             "valor": "creditos/solicitud", "ancho": 30},
            {"tipo": "check", "arg": "--generar-esquemas",
             "etiqueta": "Tomar esta corrida como baseline de esquemas",
             "valor": False},
        ],
    },
    {
        "id": "analitica",
        "nombre": "Analitica dataLayer",
        "script": "observador_analitica.py",
        "boton": "Iniciar captura",
        "parada": "centinela",
        "stop_file": STOP_FILE_ANALITICA,
        "previo": preparar_chrome,
        "ayuda": "Haces el flujo a mano; anota cada push al dataLayer y filtra el "
                 "ruido de GTM, igual que TermsAndConditionsAnalyticsSolid.",
        "campos": [
            {"tipo": "texto", "arg": "--flujo", "etiqueta": "Flujo",
             "valor": "terminos", "ancho": 26},
            {"tipo": "texto", "arg": "--solo-url", "etiqueta": "Solo URL que contenga",
             "valor": "creditos/solicitud", "ancho": 30},
        ],
    },
    {
        "id": "biometria",
        "nombre": "Suite biometria",
        "comando": [mvn, "test"],
        "cwd": FRAMEWORK,
        "boton": "Correr suite",
        # mvn no tiene parada limpia; cortarlo solo pierde la corrida de tests
        "parada": "terminar",
        "ayuda": "Corre testng-biometry.xml del framework Java con los datos que pongas.",
        "argumentos": args_biometria,
        "campos": [
            {"tipo": "texto", "solo_forma": True, "etiqueta": "Id caso",
             "valor": "", "ancho": 14, "requerido": True},
            {"tipo": "opcion", "solo_forma": True, "etiqueta": "Tipo doc",
             "opciones": ["CO1C", "CO1E"], "valor": "CO1C", "ancho": 8},
            {"tipo": "texto", "solo_forma": True, "etiqueta": "Documento",
             "valor": "", "ancho": 20, "requerido": True},
        ],
    },
    {
        "id": "bizagi_cancelar",
        "nombre": "Cancelar caso Bizagi",
        "script": "bizagi_cancel_case.py",
        "boton": "Cancelar caso",
        "parada": None,
        "ayuda": "Busca la ultima solicitud del documento y la cancela en Bizagi.",
        "confirmar": ("Se va a cancelar la ultima solicitud del documento "
                      "{Tipo doc} {Documento} en Bizagi.\n\nSeguir?"),
        "estado": estado_cancelacion,
        "campos": [
            {"tipo": "check", "arg": "--headless",
             "etiqueta": "Sin ventana del navegador", "valor": True},
            {"tipo": "opcion", "arg": None, "etiqueta": "Tipo doc",
             "opciones": ["CC", "CE"], "valor": "CC", "ancho": 6},
            {"tipo": "texto", "arg": None, "etiqueta": "Documento",
             "valor": "", "ancho": 20, "requerido": True},
        ],
    },
    {
        "id": "bizagi_consultar",
        "nombre": "Consultar caso Bizagi",
        "script": "bizagi_consultar_caso.py",
        "boton": "Consultar",
        # deja el navegador abierto a proposito; se cierra con Detener
        "parada": "terminar",
        "ayuda": "Muestra la ultima solicitud y deja el navegador abierto.",
        "estado": estado_consulta,
        "campos": [
            {"tipo": "opcion", "arg": None, "etiqueta": "Tipo doc",
             "opciones": ["CC", "CE"], "valor": "CC", "ancho": 6},
            {"tipo": "texto", "arg": None, "etiqueta": "Documento",
             "valor": "", "ancho": 20, "requerido": True},
        ],
    },
]


# --- lectura de la evidencia ----------------------------------------------
def _json_o_nada(ruta):
    try:
        with open(ruta, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError, UnicodeDecodeError):
        return None


def leer_corrida(d):
    """Una fila de la tabla, armada con lo que la corrida ya dejo en disco."""
    nombre = os.path.basename(d)
    fila = {"dir": d, "nombre": nombre, "flujo": "", "pasos": 0, "requests": 0,
            "ws": 0, "fallos": 0, "esquemas": "-", "orden": os.path.getmtime(d)}
    m = RE_MARCA.search(nombre)
    fila["fecha"] = ("%s-%s %s:%s" % (m.group(2), m.group(3), m.group(4), m.group(5))
                     if m else time.strftime("%m-%d %H:%M",
                                             time.localtime(fila["orden"])))
    resumen = _json_o_nada(os.path.join(d, "resumen.json"))
    if resumen:
        pasos = resumen.get("pasos") or []
        # --rehacer-reporte guarda el nombre de la carpeta como flujo, con
        # marca de tiempo incluida; para la tabla sobra
        fila["flujo"] = RE_MARCA.sub("", resumen.get("flujo", ""))
        fila["pasos"] = len(pasos)
        fila["requests"] = sum(p.get("requests", 0) for p in pasos)
        fila["ws"] = sum(p.get("websocket", 0) for p in pasos)
        fila["fallos"] = sum(p.get("fallos", 0) for p in pasos)
    if not fila["flujo"]:
        fila["flujo"] = RE_MARCA.sub("", nombre)
    val = _json_o_nada(os.path.join(d, "validacion_esquemas.json"))
    if val:
        n = (val.get("resumen") or {}).get("con_hallazgos", 0)
        fila["esquemas"] = "ok" if not n else "%d hallazgo%s" % (n, "" if n == 1 else "s")
    return fila


def listar_corridas():
    if not os.path.isdir(EVIDENCIAS):
        return []
    dirs = [os.path.join(EVIDENCIAS, d) for d in os.listdir(EVIDENCIAS)
            if os.path.isdir(os.path.join(EVIDENCIAS, d))]
    return sorted((leer_corrida(d) for d in dirs),
                  key=lambda x: x["orden"], reverse=True)


# --- una pestana por herramienta ------------------------------------------
class Herramienta(ttk.Frame):
    def __init__(self, padre, spec):
        super().__init__(padre)
        self.spec = spec
        self.proc = None
        self.cola = queue.Queue()
        self.salida = []          # todo el stdout, para el veredicto final
        self.dir_corrida = None   # carpeta de evidencia de la ultima corrida
        self.vars = {}
        self._construir()

    # -- UI
    def _construir(self):
        cfg = ttk.Frame(self, padding=(12, 12, 12, 4))
        cfg.pack(fill="x")
        if self.spec.get("ayuda"):
            ttk.Label(cfg, text=self.spec["ayuda"], foreground="#666").grid(
                row=0, column=0, columnspan=8, sticky="w", pady=(0, 10))

        col = 0
        fila = 1
        for campo in self.spec["campos"]:
            etq = campo["etiqueta"]
            if campo["tipo"] == "check":
                var = tk.BooleanVar(value=campo["valor"])
                ttk.Checkbutton(cfg, text=etq, variable=var).grid(
                    row=99, column=0, columnspan=8, sticky="w", pady=(8, 0))
            else:
                var = tk.StringVar(value=campo["valor"])
                ttk.Label(cfg, text=etq).grid(row=fila, column=col, sticky="w")
                if campo["tipo"] == "opcion":
                    w = ttk.Combobox(cfg, textvariable=var, values=campo["opciones"],
                                     width=campo.get("ancho", 10), state="readonly")
                else:
                    w = ttk.Entry(cfg, textvariable=var, width=campo.get("ancho", 24))
                    w.bind("<Return>", lambda _e: self.ejecutar())
                w.grid(row=fila, column=col + 1, padx=(6, 18), sticky="w")
                col += 2
                if col >= 6:
                    col, fila = 0, fila + 1
            self.vars[etq] = var

        botones = ttk.Frame(self, padding=(12, 10, 12, 4))
        botones.pack(fill="x")
        self.b_ir = ttk.Button(botones, text=self.spec["boton"], command=self.ejecutar)
        self.b_ir.pack(side="left")
        self.b_parar = ttk.Button(botones, text=self._texto_parada(),
                                  command=self.detener, state="disabled")
        self.b_parar.pack(side="left", padx=8)
        self.b_reporte = None
        if self.spec["parada"] == "centinela":
            # separadas a proposito: a veces uno corta la captura para seguir
            # navegando y solo quiere el reporte mas tarde
            self.b_solo_parar = ttk.Button(botones, text="Detener sin reporte",
                                           command=self.detener_seco,
                                           state="disabled")
            self.b_solo_parar.pack(side="left")
            self.b_reporte = ttk.Button(botones, text="Generar reporte",
                                        command=self.generar_reporte)
            self.b_reporte.pack(side="left", padx=8)
        else:
            self.b_solo_parar = None
        ttk.Button(botones, text="Limpiar log", command=self.limpiar).pack(side="left")

        self.estado = tk.StringVar(value="Listo.")
        self.lbl_estado = ttk.Label(self, textvariable=self.estado, padding=(14, 2))
        self.lbl_estado.pack(fill="x")

        marco = ttk.Frame(self, padding=(12, 4, 12, 12))
        marco.pack(fill="both", expand=True)
        self.log = tk.Text(marco, wrap="none", bg="#101010", fg="#d8d8d8",
                           insertbackground="#d8d8d8", font=("Consolas", 9),
                           relief="flat")
        barra = ttk.Scrollbar(marco, orient="vertical", command=self.log.yview)
        self.log.configure(yscrollcommand=barra.set, state="disabled")
        self.log.pack(side="left", fill="both", expand=True)
        barra.pack(side="right", fill="y")
        for tag, color in (("paso", "#7fb2ff"), ("ws", "#c9a0ff"), ("mal", "#ff8a80"),
                           ("bien", "#9be29b"), ("panel", "#ffd479")):
            self.log.tag_configure(tag, foreground=color)

    def _texto_parada(self):
        return {"centinela": "Detener y generar reporte",
                "terminar": "Cerrar navegador"}.get(self.spec["parada"], "Detener")

    # -- log
    def emitir(self, linea, tag=None):
        self.cola.put((linea, tag))

    def escribir(self, linea, tag=None):
        if tag is None:
            t = linea.lstrip()
            # el fallo manda sobre el tipo de linea: un [ws ...] FAIL tiene
            # que verse rojo, no del morado de los websockets
            if "FAIL" in linea or "ERROR" in linea or t.startswith("!"):
                tag = "mal"
            elif t.startswith("[paso"):
                tag = "paso"
            elif t.startswith("[ws"):
                tag = "ws"
            elif t.startswith("ok ") or "Reporte:" in linea:
                tag = "bien"
        self.log.configure(state="normal")
        self.log.insert("end", linea.rstrip() + "\n", tag or ())
        self.log.see("end")
        self.log.configure(state="disabled")

    def drenar(self):
        try:
            while True:
                self.escribir(*self.cola.get_nowait())
        except queue.Empty:
            pass

    def limpiar(self):
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

    # -- ejecucion
    def ocupada(self):
        return self.proc is not None and self.proc.poll() is None

    def _base(self):
        """El comando sin argumentos: un script de este repo, o algo externo."""
        if self.spec.get("comando"):
            return [c() if callable(c) else c for c in self.spec["comando"]]
        return [sys.executable, "-u", os.path.join(AQUI, self.spec["script"])]

    def valores(self):
        return {c["etiqueta"]: str(self.vars[c["etiqueta"]].get()).strip()
                for c in self.spec["campos"] if c["tipo"] != "check"}

    def construir_cmd(self):
        """Del formulario al comando. Los posicionales van al final, en orden."""
        cmd = self._base()
        posicionales = []
        for campo in self.spec["campos"]:
            valor = self.vars[campo["etiqueta"]].get()
            if campo["tipo"] == "check":
                if valor:
                    cmd.append(campo["arg"])
                continue
            valor = str(valor).strip()
            if not valor:
                if campo.get("requerido"):
                    return None, campo["etiqueta"]
                continue
            if campo.get("solo_forma"):
                continue          # alimenta el hook de argumentos, no la linea
            if campo.get("arg"):
                cmd += [campo["arg"], valor]
            else:
                posicionales.append(valor)
        if self.spec["parada"] == "centinela":
            cmd += ["--stop-file", self.spec["stop_file"]]
        return cmd + posicionales, None

    def ejecutar(self):
        if self.ocupada():
            return
        cmd, falta = self.construir_cmd()
        if falta:
            messagebox.showinfo("Falta un dato", "Escribe %s." % falta)
            return
        aviso = self.spec.get("confirmar")
        if aviso:
            texto = aviso
            for etq, var in self.vars.items():
                texto = texto.replace("{%s}" % etq, str(var.get()))
            if not messagebox.askyesno(self.spec["nombre"], texto):
                return
        self.limpiar()
        self.salida = []
        self.b_ir.configure(state="disabled")
        self.lbl_estado.configure(foreground="")
        self.estado.set("Ejecutando...")
        threading.Thread(target=self._hilo, args=(cmd,), daemon=True).start()

    def _hilo(self, cmd, usar_previo=True):
        """Pre-flight, subproceso y bombeo del stdout. Fuera del hilo de la UI
        para que la ventana no se congele.

        usar_previo=False para comandos del mismo script que no necesitan el
        pre-flight (--rehacer-reporte trabaja sobre disco, no sobre Chrome).
        """
        previo = self.spec.get("previo") if usar_previo else None
        if previo and not previo(self.emitir):
            self.after(0, self._fin)
            return
        hook = self.spec.get("argumentos")
        if hook:
            extra = hook(self.valores(), self.emitir)
            if extra is None:
                self.after(0, self._fin)
                return
            cmd = cmd + extra
        stop = self.spec.get("stop_file")
        if stop and os.path.exists(stop):
            os.remove(stop)                    # sobra de una corrida anterior
        self.emitir("> " + " ".join(cmd[1:]), "panel")
        # el hijo imprime en utf-8 y no en la codificacion de la consola, para
        # que las tildes no lleguen rotas al log
        entorno = dict(os.environ, PYTHONIOENCODING="utf-8")
        try:
            self.proc = subprocess.Popen(
                cmd, cwd=self.spec.get("cwd", AQUI),
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                stdin=subprocess.DEVNULL, text=True, encoding="utf-8",
                errors="replace", bufsize=1, env=entorno)
        except OSError as e:
            self.emitir("! No pude lanzar el proceso: %s" % e, "mal")
            self.after(0, self._fin)
            return
        self.after(0, self._corriendo)
        for linea in self.proc.stdout:
            self.salida.append(linea)
            if linea.startswith("Evidencia:"):
                self.dir_corrida = linea.split(":", 1)[1].strip()
            self.cola.put((linea, None))
        self.proc.wait()
        self.after(0, self._fin)

    def _corriendo(self):
        if self.spec["parada"]:
            self.b_parar.configure(state="normal")
        if self.b_solo_parar:
            self.b_solo_parar.configure(state="normal")
        self.estado.set({"centinela": "Capturando. Navega en Chrome; al terminar, Detener.",
                         "terminar": "Corriendo. El navegador queda abierto."}
                        .get(self.spec["parada"], "Corriendo..."))

    def _fin(self):
        self.b_ir.configure(state="normal")
        self.b_parar.configure(state="disabled")
        if self.b_solo_parar:
            self.b_solo_parar.configure(state="disabled")
        self.proc = None
        veredicto = self.spec.get("estado")
        if veredicto:
            tag, msg = veredicto("".join(self.salida))
            self.estado.set(msg)
            self.lbl_estado.configure(
                foreground={"bien": "#1b7f2b", "mal": "#c62828"}.get(tag, ""))
        else:
            self.estado.set("Terminado.")
        for cb in getattr(self, "al_terminar", []):
            cb()

    def detener(self, con_reporte=True):
        if not self.ocupada():
            return
        self.b_parar.configure(state="disabled")
        if self.b_solo_parar:
            self.b_solo_parar.configure(state="disabled")
        if self.spec["parada"] == "centinela":
            self.estado.set("Cerrando y generando reporte..." if con_reporte
                            else "Cerrando sin generar el reporte...")
            with open(self.spec["stop_file"], "w", encoding="utf-8") as f:
                f.write("" if con_reporte else SIN_REPORTE)
        else:
            self.estado.set("Cerrando...")
            self.proc.terminate()

    def detener_seco(self):
        """Corta la captura y deja la evidencia cruda; el reporte, cuando quieras."""
        self.detener(con_reporte=False)

    def generar_reporte(self):
        """--rehacer-reporte sobre la ultima corrida (o la mas reciente en disco)."""
        if self.ocupada():
            messagebox.showinfo("Captura en curso",
                                "Detén la captura antes de generar el reporte.")
            return
        d = self.dir_corrida
        if not d or not os.path.isdir(d):
            corridas = listar_corridas()
            d = corridas[0]["dir"] if corridas else None
        if not d:
            messagebox.showinfo("Sin corridas", "Todavia no hay evidencia capturada.")
            return
        self.limpiar()
        self.salida = []
        self.dir_corrida = d
        self.estado.set("Generando reporte de %s..." % os.path.basename(d))
        cmd = [sys.executable, "-u", os.path.join(AQUI, self.spec["script"]),
               "--rehacer-reporte", d]
        threading.Thread(target=self._hilo, args=(cmd, False), daemon=True).start()


# --- pestana de corridas ---------------------------------------------------
class Corridas(ttk.Frame):
    def __init__(self, padre, panel):
        super().__init__(padre)
        self.panel = panel
        self.filas = {}
        self._construir()
        self.refrescar()

    def _construir(self):
        cols = ("fecha", "flujo", "pasos", "requests", "ws", "fallos", "esquemas")
        anchos = (110, 190, 55, 80, 50, 65, 110)
        titulos = ("Fecha", "Flujo", "Pasos", "Requests", "WS", "4xx/5xx", "Esquemas")
        marco = ttk.Frame(self, padding=12)
        marco.pack(fill="both", expand=True)
        self.tabla = ttk.Treeview(marco, columns=cols, show="headings", selectmode="browse")
        for c, a, t in zip(cols, anchos, titulos):
            self.tabla.heading(c, text=t)
            self.tabla.column(c, width=a,
                              anchor="w" if c in ("flujo", "fecha", "esquemas") else "center")
        barra = ttk.Scrollbar(marco, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=barra.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        barra.pack(side="right", fill="y")
        self.tabla.tag_configure("mal", foreground="#c62828")
        self.tabla.bind("<Double-1>", lambda _e: self.abrir_reporte())

        pie = ttk.Frame(self, padding=(12, 0, 12, 6))
        pie.pack(fill="x")
        ttk.Button(pie, text="Abrir reporte", command=self.abrir_reporte).pack(side="left")
        ttk.Button(pie, text="Abrir carpeta", command=self.abrir_carpeta).pack(side="left", padx=8)
        ttk.Button(pie, text="Revalidar", command=self.revalidar).pack(side="left")
        ttk.Button(pie, text="Refrescar", command=self.refrescar).pack(side="left", padx=8)
        ttk.Button(pie, text="Borrar", command=self.borrar).pack(side="right")

        self.estado = tk.StringVar(value="")
        ttk.Label(self, textvariable=self.estado, padding=(14, 0, 14, 8)).pack(fill="x")

    def refrescar(self):
        self.tabla.delete(*self.tabla.get_children())
        self.filas.clear()
        corridas = listar_corridas()
        for c in corridas:
            tag = "mal" if (c["fallos"] or "hallazgo" in c["esquemas"]) else ""
            iid = self.tabla.insert(
                "", "end", tags=(tag,),
                values=(c["fecha"], c["flujo"], c["pasos"], c["requests"],
                        c["ws"] or "", c["fallos"] or "", c["esquemas"]))
            self.filas[iid] = c
        self.estado.set("%d corrida(s) en evidences/" % len(corridas))

    def _seleccionada(self):
        sel = self.tabla.selection()
        if not sel:
            messagebox.showinfo("Sin seleccion", "Escoge una corrida de la lista.")
            return None
        return self.filas.get(sel[0])

    def abrir_reporte(self):
        c = self._seleccionada()
        if not c:
            return
        ruta = os.path.join(c["dir"], "reporte.html")
        if not os.path.exists(ruta):
            messagebox.showinfo("Sin reporte", "Esa corrida no tiene reporte.html.")
            return
        webbrowser.open("file:///" + os.path.abspath(ruta).replace("\\", "/"))

    def abrir_carpeta(self):
        c = self._seleccionada()
        if not c:
            return
        if hasattr(os, "startfile"):
            os.startfile(c["dir"])          # Windows
        else:
            webbrowser.open("file:///" + c["dir"].replace("\\", "/"))

    def revalidar(self):
        """--rehacer-reporte sobre la corrida: es lo que hay que correr cada vez
        que se corrige el baseline de esquemas."""
        c = self._seleccionada()
        if not c:
            return
        herr = self.panel.herramientas["observador"]
        if herr.ocupada():
            messagebox.showinfo("Ocupado", "Hay una captura en curso.")
            return
        self.panel.tabs.select(herr)
        herr.limpiar()
        herr.salida = []
        herr.estado.set("Revalidando %s..." % c["nombre"])
        cmd = [sys.executable, "-u", os.path.join(AQUI, "observador_flujo.py"),
               "--rehacer-reporte", c["dir"]]
        threading.Thread(target=herr._hilo, args=(cmd, False), daemon=True).start()

    def borrar(self):
        c = self._seleccionada()
        if not c:
            return
        if not messagebox.askyesno(
                "Borrar evidencia",
                "Se va a borrar de forma permanente:\n\n%s\n\n"
                "Con sus pantallazos, requests y reporte. No hay papelera.\n\n"
                "Seguro?" % c["dir"]):
            return
        try:
            shutil.rmtree(c["dir"])
        except OSError as e:
            messagebox.showerror("No se pudo borrar", str(e))
            return
        self.refrescar()


# --- validaciones API por lote --------------------------------------------
TIPOS_DOC = ["CC", "CE"]


def solo_unicos(docs):
    """El DataProvider deduplica por numero con un LinkedHashSet; se hace lo
    mismo aqui para poder avisar cuantos van a correr de verdad."""
    vistos, unicos = set(), []
    for tipo, numero in docs:
        if numero not in vistos:
            vistos.add(numero)
            unicos.append((tipo, numero))
    return unicos


class Validaciones(ttk.Frame):
    """Corre ApiTest.testValidationServices sobre una lista de documentos.

    Pega la lista, se reescribe el array raw del DataProvider, se compila y se
    corre. El log de Maven son miles de lineas: aqui solo se muestran las que
    sirven y el resto queda en un archivo aparte.
    """

    def __init__(self, padre):
        super().__init__(padre)
        self.proc = None
        self.cola = queue.Queue()
        self.servicios = list(SERVICIOS_FIJOS)   # columnas, en orden del Excel
        self.resultados = {}         # documento -> {servicio: estado}
        self.ruta_log = None
        self.fallo_build = False
        self._construir()

    def _construir(self):
        arriba = ttk.Frame(self, padding=(12, 12, 12, 4))
        arriba.pack(fill="x")
        ttk.Label(arriba, text="Tipo").grid(row=0, column=0, sticky="w")
        ttk.Label(arriba, text="Identificacion").grid(row=0, column=1, sticky="w",
                                                      padx=(6, 0))
        self.v_tipo = tk.StringVar(value=TIPOS_DOC[0])
        ttk.Combobox(arriba, textvariable=self.v_tipo, values=TIPOS_DOC, width=5,
                     state="readonly").grid(row=1, column=0, sticky="w")
        self.v_doc = tk.StringVar()
        e = ttk.Entry(arriba, textvariable=self.v_doc, width=20)
        e.grid(row=1, column=1, sticky="w", padx=(6, 12))
        e.bind("<Return>", lambda _ev: self.agregar())     # numero, Enter, numero...
        self.foco = e
        ttk.Button(arriba, text="Agregar", command=self.agregar).grid(row=1, column=2)
        ttk.Button(arriba, text="Quitar", command=self.quitar).grid(row=1, column=3,
                                                                    padx=8)
        ttk.Button(arriba, text="Cargar lista actual",
                   command=self.cargar_actual).grid(row=1, column=4)
        ttk.Button(arriba, text="Vaciar", command=self.vaciar).grid(row=1, column=5,
                                                                    padx=8)

        md = ttk.Frame(self, padding=(12, 8, 12, 0))
        md.pack(fill="x")
        self.docs = []
        self.tabla_docs = ttk.Treeview(md, columns=("tipo", "doc"), show="headings",
                                       selectmode="extended", height=6)
        self.tabla_docs.heading("tipo", text="Tipo")
        self.tabla_docs.heading("doc", text="Identificacion")
        self.tabla_docs.column("tipo", width=60, anchor="center")
        self.tabla_docs.column("doc", width=180, anchor="w")
        sb = ttk.Scrollbar(md, orient="vertical", command=self.tabla_docs.yview)
        self.tabla_docs.configure(yscrollcommand=sb.set)
        self.tabla_docs.pack(side="left", fill="x", expand=True)
        sb.pack(side="right", fill="y")
        self.tabla_docs.tag_configure("repetido", foreground="#999")
        self.tabla_docs.bind("<Delete>", lambda _ev: self.quitar())

        botones = ttk.Frame(self, padding=(12, 8, 12, 4))
        botones.pack(fill="x")
        self.b_ir = ttk.Button(botones, text="Reemplazar y ejecutar", command=self.ejecutar)
        self.b_ir.pack(side="left")
        self.b_parar = ttk.Button(botones, text="Detener", command=self.detener,
                                  state="disabled")
        self.b_parar.pack(side="left", padx=8)
        ttk.Button(botones, text="Cargar lista actual",
                   command=self.cargar_actual).pack(side="left")
        ttk.Button(botones, text="Abrir log completo",
                   command=self.abrir_log).pack(side="left", padx=8)
        self.b_excel = ttk.Button(botones, text="Exportar a Excel",
                                  command=self.exportar, state="disabled")
        self.b_excel.pack(side="right")

        self.estado = tk.StringVar(value="Pega la lista de documentos.")
        self.lbl = ttk.Label(self, textvariable=self.estado, padding=(14, 2))
        self.lbl.pack(fill="x")

        panes = ttk.PanedWindow(self, orient="vertical")
        panes.pack(fill="both", expand=True, padx=12, pady=(4, 12))
        mt = ttk.Frame(panes)
        self.tabla = ttk.Treeview(mt, columns=("documento",), show="headings",
                                  selectmode="browse", height=8)
        self.tabla.heading("documento", text="Documento")
        self.tabla.column("documento", width=130, anchor="w")
        sb2 = ttk.Scrollbar(mt, orient="vertical", command=self.tabla.yview)
        sbh = ttk.Scrollbar(mt, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(yscrollcommand=sb2.set, xscrollcommand=sbh.set)
        sb2.pack(side="right", fill="y")
        sbh.pack(side="bottom", fill="x")
        self.tabla.pack(side="left", fill="both", expand=True)
        self.tabla.tag_configure("mal", foreground="#c62828")
        self.tabla.tag_configure("bien", foreground="#1b7f2b")
        panes.add(mt, weight=3)

        ml = ttk.Frame(panes)
        self.log = tk.Text(ml, wrap="none", bg="#101010", fg="#d8d8d8", height=8,
                           font=("Consolas", 9), relief="flat")
        sb3 = ttk.Scrollbar(ml, orient="vertical", command=self.log.yview)
        self.log.configure(yscrollcommand=sb3.set, state="disabled")
        self.log.pack(side="left", fill="both", expand=True)
        sb3.pack(side="right", fill="y")
        for tag, color in (("mal", "#ff8a80"), ("bien", "#9be29b"), ("panel", "#ffd479")):
            self.log.tag_configure(tag, foreground=color)
        panes.add(ml, weight=2)

        self._columnas()      # las columnas fijas se ven desde el arranque

    # -- log
    def emitir(self, linea, tag=None):
        self.cola.put((linea, tag))

    def drenar(self):
        try:
            while True:
                linea, tag = self.cola.get_nowait()
                self.log.configure(state="normal")
                self.log.insert("end", linea.rstrip() + "\n", tag or ())
                self.log.see("end")
                self.log.configure(state="disabled")
        except queue.Empty:
            pass

    # -- lista de documentos
    def agregar(self):
        numero = re.sub(r"\D", "", self.v_doc.get())
        if not numero:
            return
        self.docs.append((self.v_tipo.get(), numero))
        self.v_doc.set("")
        self.foco.focus_set()
        self.refrescar_docs()

    def quitar(self):
        for iid in sorted(self.tabla_docs.selection(), key=int, reverse=True):
            self.docs.pop(int(iid))
        self.refrescar_docs()

    def vaciar(self):
        self.docs = []
        self.refrescar_docs()

    def refrescar_docs(self):
        self.tabla_docs.delete(*self.tabla_docs.get_children())
        vistos = set()
        for i, (tipo, numero) in enumerate(self.docs):
            repe = numero in vistos
            vistos.add(numero)
            self.tabla_docs.insert("", "end", iid=str(i),
                                   tags=("repetido",) if repe else (),
                                   values=(tipo, numero))
        unicos = solo_unicos(self.docs)
        rep = len(self.docs) - len(unicos)
        if not self.docs:
            self.estado.set("Agrega documentos a la lista.")
        else:
            self.estado.set("%d documento(s), %d se van a correr%s"
                            % (len(self.docs), len(unicos),
                               "  (%d repetido(s), en gris)" % rep if rep else ""))
        self.tabla_docs.see(str(len(self.docs) - 1)) if self.docs else None

    def cargar_actual(self):
        """Trae al cuadro la lista que hoy tiene el DataProvider."""
        ruta = os.path.join(FRAMEWORK, DATA_PROVIDER)
        try:
            with open(ruta, encoding="utf-8") as f:
                bloque = RE_RAW.search(f.read())
        except OSError as e:
            messagebox.showerror("No se pudo leer", str(e))
            return
        if not bloque:
            messagebox.showerror("No encontrado",
                                 "No hallé el array raw en %s" % DATA_PROVIDER)
            return
        self.docs = re.findall(r'\{"(\w+)",\s*"(\d+)"\}', bloque.group(0))
        self.refrescar_docs()

    def escribir_provider(self, unicos):
        """Reescribe el array raw. Guarda copia del original antes de tocarlo."""
        ruta = os.path.join(FRAMEWORK, DATA_PROVIDER)
        with open(ruta, encoding="utf-8") as f:
            contenido = f.read()
        if not RE_RAW.search(contenido):
            self.emitir("! No encontre el array raw en %s" % DATA_PROVIDER, "mal")
            return False
        os.makedirs(BACKUPS, exist_ok=True)
        shutil.copy2(ruta, os.path.join(
            BACKUPS, "DataProviderUtil_%s.java" % time.strftime("%Y%m%d_%H%M%S")))
        filas = "\n".join('                {"%s", "%s"},' % (t, n) for t, n in unicos)
        nuevo = "        Object[][] raw = {\n%s\n        };" % filas
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(RE_RAW.sub(lambda _m: nuevo, contenido, count=1))
        self.emitir("> DataProvider reescrito con %d documento(s). Copia del "
                    "original en backups/." % len(unicos), "panel")
        return True

    # -- ejecucion
    def ocupada(self):
        return self.proc is not None and self.proc.poll() is None

    def ejecutar(self):
        if self.ocupada():
            return
        unicos = solo_unicos(self.docs)
        if not unicos:
            messagebox.showinfo("Falta la lista",
                                "Agrega al menos un documento a la lista.")
            return
        if not os.path.isdir(FRAMEWORK):
            messagebox.showerror("Sin framework", "No encuentro %s" % FRAMEWORK)
            return
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")
        self.tabla.delete(*self.tabla.get_children())
        self.servicios = list(SERVICIOS_FIJOS)
        self.resultados, self.fallo_build = {}, False
        self._columnas()
        self.b_ir.configure(state="disabled")
        self.b_excel.configure(state="disabled")
        self.lbl.configure(foreground="")
        self.estado.set("Compilando y ejecutando...")
        threading.Thread(target=self._hilo, args=(unicos,), daemon=True).start()

    def _hilo(self, unicos):
        try:
            if not self.escribir_provider(unicos):
                self.after(0, self._fin)
                return
        except OSError as e:
            self.emitir("! No pude escribir el DataProvider: %s" % e, "mal")
            self.after(0, self._fin)
            return

        os.makedirs(LOGS, exist_ok=True)
        self.ruta_log = os.path.join(
            LOGS, "validaciones_%s.log" % time.strftime("%Y%m%d_%H%M%S"))
        # test-compile antes de surefire:test: sin eso Maven corre las clases
        # ya compiladas y ejecuta la lista vieja
        cmd = [mvn(), "-o", "test-compile", "surefire:test"]
        self.emitir("> " + " ".join(cmd), "panel")
        self.emitir("> log completo: %s" % self.ruta_log, "panel")
        entorno = dict(os.environ, PYTHONIOENCODING="utf-8")
        try:
            self.proc = subprocess.Popen(
                cmd, cwd=FRAMEWORK, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                stdin=subprocess.DEVNULL, text=True, encoding="utf-8",
                errors="replace", bufsize=1, env=entorno)
        except OSError as e:
            self.emitir("! No pude lanzar Maven: %s" % e, "mal")
            self.after(0, self._fin)
            return
        self.after(0, lambda: self.b_parar.configure(state="normal"))
        with open(self.ruta_log, "w", encoding="utf-8") as archivo:
            for linea in self.proc.stdout:
                archivo.write(linea)
                self._linea(linea)
        self.proc.wait()
        self.after(0, self._fin)

    def _linea(self, linea):
        """Del torrente de Maven solo sube al panel lo que se mira."""
        m = RE_EXCEL.search(linea)
        if m:
            doc, servicio, estado = m.group(1), m.group(2).strip(), m.group(3).strip()
            if servicio.lower() in SERVICIOS_OCULTOS:
                return                    # sigue en el log completo del disco
            self.after(0, self._resultado, doc, servicio, estado)
            malo = estado.upper().startswith("FAIL")
            self.emitir("  %-12s %-26s %s" % (doc, servicio, estado),
                        "mal" if malo else "bien")
            return
        t = linea.strip()
        if t.startswith("Tests run:") or "BUILD " in t:
            if "BUILD FAILURE" in t:
                self.fallo_build = True
            self.emitir(t, "panel")

    def _columnas(self):
        cols = ["documento"] + self.servicios
        self.tabla.configure(columns=cols)
        self.tabla.heading("documento", text="Documento")
        self.tabla.column("documento", width=130, anchor="w")
        for s in self.servicios:
            self.tabla.heading(s, text=s)
            self.tabla.column(s, width=190, anchor="w")

    def _resultado(self, doc, servicio, estado):
        if servicio not in self.servicios:
            self.servicios.append(servicio)
            self.servicios.sort(key=_orden_servicio)
            self._columnas()
        self.resultados.setdefault(doc, {})[servicio] = estado
        self._pintar()

    def _pintar(self):
        self.tabla.delete(*self.tabla.get_children())
        for doc, servs in self.resultados.items():
            valores = [doc] + [servs.get(s, "") for s in self.servicios]
            estados = " ".join(servs.values()).upper()
            tag = "mal" if "FAIL" in estados else ("bien" if "PASS" in estados else "")
            self.tabla.insert("", "end", tags=(tag,), values=valores)

    def _fin(self):
        self.b_ir.configure(state="normal")
        self.b_parar.configure(state="disabled")
        self.proc = None
        docs = len(self.resultados)
        malos = sum(1 for s in self.resultados.values()
                    if "FAIL" in " ".join(s.values()).upper())
        if docs:
            self.b_excel.configure(state="normal")
            self.estado.set("%d documento(s) con resultado, %d con algun FAIL."
                            % (docs, malos))
            self.lbl.configure(foreground="#c62828" if malos else "#1b7f2b")
            if self.fallo_build:
                # ~11 failures fijos de otras clases que piden parametros de
                # testng.xml; no dicen nada de fillDataApi
                self.emitir("> BUILD FAILURE es normal aqui: son los tests de otras "
                            "clases que piden parametros de testng.xml.", "panel")
        else:
            self.estado.set("No hubo lineas de resultado. Revisa el log completo.")
            self.lbl.configure(foreground="#c62828")

    def detener(self):
        if self.ocupada():
            self.b_parar.configure(state="disabled")
            self.proc.terminate()

    def abrir_log(self):
        if not self.ruta_log or not os.path.exists(self.ruta_log):
            messagebox.showinfo("Sin log", "Todavia no hay log de una corrida.")
            return
        if hasattr(os, "startfile"):
            os.startfile(self.ruta_log)
        else:
            webbrowser.open("file:///" + self.ruta_log.replace("\\", "/"))

    def exportar(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror("Falta openpyxl", "pip install openpyxl")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Validaciones"
        ws.append(["Documento"] + self.servicios)
        for c in ws[1]:
            c.font = Font(bold=True)
        verde = PatternFill("solid", start_color="C6EFCE")
        rojo = PatternFill("solid", start_color="FFC7CE")
        for doc, servs in self.resultados.items():
            ws.append([doc] + [servs.get(s, "") for s in self.servicios])
            for c in ws[ws.max_row][1:]:
                v = str(c.value or "").upper()
                if v.startswith("FAIL"):
                    c.fill = rojo
                elif v.startswith("PASS"):
                    c.fill = verde
        ws.column_dimensions["A"].width = 16
        for i in range(len(self.servicios)):
            ws.column_dimensions[chr(66 + i)].width = 34
        os.makedirs(LOGS, exist_ok=True)
        ruta = os.path.join(LOGS, "validaciones_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S"))
        try:
            wb.save(ruta)
        except OSError as e:
            messagebox.showerror("No se pudo guardar", str(e))
            return
        self.estado.set("Excel: %s" % ruta)
        if hasattr(os, "startfile"):
            os.startfile(ruta)


# --- usuarios de prueba ----------------------------------------------------
class Usuarios(ttk.Frame):
    """Libreta de usuarios de prueba: quien es, para que sirve y como quedo.

    Se guarda en usuarios_prueba.json apenas se toca algo, para que cerrar el
    panel de golpe no pierda nada.
    """

    COLS = ("tipo", "usuario", "clave", "descripcion", "status", "falla")
    TITULOS = ("Tipo", "Usuario", "Contrasena", "Descripcion", "Status", "Falla")
    ANCHOS = (55, 125, 105, 245, 105, 215)
    ESTADOS = ["Disponible", "En uso", "Bloqueado", "Quemado", "Con falla"]
    # CC/CE es como los piden los scripts de Bizagi; los servicios usan
    # CO1C/CO1E, que es otra codificacion del mismo tipo de documento
    TIPOS = ["CC", "CE"]

    def __init__(self, padre):
        super().__init__(padre)
        self.datos = []
        self.editando = None      # indice que se esta editando, o None
        self.vars = {}
        self._construir()
        self.cargar()

    def _construir(self):
        form = ttk.Frame(self, padding=(12, 12, 12, 4))
        form.pack(fill="x")
        anchos = {"usuario": 18, "clave": 16, "descripcion": 34, "falla": 34}
        col = 0
        for clave, titulo in zip(self.COLS, self.TITULOS):
            ttk.Label(form, text=titulo).grid(row=0, column=col, sticky="w", padx=(0, 6))
            var = tk.StringVar()
            if clave == "tipo":
                var.set(self.TIPOS[0])
                w = ttk.Combobox(form, textvariable=var, values=self.TIPOS,
                                 width=5, state="readonly")
            elif clave == "status":
                w = ttk.Combobox(form, textvariable=var, values=self.ESTADOS, width=14)
            else:
                w = ttk.Entry(form, textvariable=var, width=anchos.get(clave, 18))
                w.bind("<Return>", lambda _e: self.guardar())
            w.grid(row=1, column=col, sticky="w", padx=(0, 14))
            self.vars[clave] = var
            col += 1
        self.ver_claves = tk.BooleanVar(value=False)
        ttk.Checkbutton(form, text="Mostrar contrasenas en la tabla",
                        variable=self.ver_claves,
                        command=self.refrescar).grid(row=2, column=0, columnspan=3,
                                                     sticky="w", pady=(10, 0))
        # apagado por defecto: publicar deja las claves en el historial de git
        # para siempre, y eso no se deshace borrando el archivo despues
        self.publicar_claves = tk.BooleanVar(value=False)
        ttk.Checkbutton(form, text="Publicar tambien las contrasenas",
                        variable=self.publicar_claves).grid(
                            row=2, column=3, columnspan=3, sticky="w", pady=(10, 0))

        botones = ttk.Frame(self, padding=(12, 6))
        botones.pack(fill="x")
        self.b_guardar = ttk.Button(botones, text="Agregar", command=self.guardar)
        self.b_guardar.pack(side="left")
        ttk.Button(botones, text="Limpiar campos",
                   command=self.limpiar).pack(side="left", padx=8)
        ttk.Button(botones, text="Copiar usuario",
                   command=lambda: self.copiar("usuario")).pack(side="left")
        ttk.Button(botones, text="Copiar contrasena",
                   command=lambda: self.copiar("clave")).pack(side="left", padx=8)
        ttk.Button(botones, text="Traer del repo",
                   command=self.traer).pack(side="left", padx=(16, 0))
        ttk.Button(botones, text="Publicar al repo",
                   command=self.publicar).pack(side="left", padx=8)
        ttk.Button(botones, text="Abrir carpeta de datos",
                   command=self.abrir_datos).pack(side="left")
        ttk.Button(botones, text="Eliminar", command=self.eliminar).pack(side="right")

        marco = ttk.Frame(self, padding=(12, 4, 12, 6))
        marco.pack(fill="both", expand=True)
        self.tabla = ttk.Treeview(marco, columns=self.COLS, show="headings",
                                  selectmode="browse")
        for c, a, t in zip(self.COLS, self.ANCHOS, self.TITULOS):
            self.tabla.heading(c, text=t)
            self.tabla.column(c, width=a,
                              anchor="center" if c in ("tipo", "status") else "w")
        barra = ttk.Scrollbar(marco, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=barra.set)
        self.tabla.pack(side="left", fill="both", expand=True)
        barra.pack(side="right", fill="y")
        self.tabla.tag_configure("mal", foreground="#c62828")
        self.tabla.tag_configure("libre", foreground="#1b7f2b")
        self.tabla.bind("<<TreeviewSelect>>", lambda _e: self.cargar_en_form())

        self.estado = tk.StringVar(value="")
        ttk.Label(self, textvariable=self.estado,
                  padding=(14, 0, 14, 8)).pack(fill="x")

    # -- persistencia
    @staticmethod
    def _clave(u):
        return (u.get("tipo", "").upper(), u.get("usuario", "").strip())

    def cargar(self):
        os.makedirs(DIR_DATOS, exist_ok=True)
        # el archivo vivia dentro del repo; si sigue ahi, se trae una sola vez
        if not os.path.exists(USUARIOS) and os.path.exists(USUARIOS_LEGADO):
            shutil.copy2(USUARIOS_LEGADO, USUARIOS)
        # primer arranque tras clonar: el catalogo del repo siembra el local
        if not os.path.exists(USUARIOS) and os.path.exists(USUARIOS_COMPARTIDOS):
            shutil.copy2(USUARIOS_COMPARTIDOS, USUARIOS)
        datos = _json_o_nada(USUARIOS)
        self.datos = datos if isinstance(datos, list) else []
        self.refrescar()

    def volcar(self):
        """Guarda copia de lo anterior y escribe de forma atomica.

        La copia es contra el error humano (borrar la fila que no era); el
        os.replace es contra el corte a media escritura, que dejaria el JSON
        partido y sin poder leerse.
        """
        try:
            os.makedirs(BACKUPS, exist_ok=True)
            if os.path.exists(USUARIOS):
                # con precision de segundo, varios guardados seguidos se
                # pisaban entre si y quedaba una sola copia
                marca = "%s_%03d" % (time.strftime("%Y%m%d_%H%M%S"),
                                     int(time.time() * 1000) % 1000)
                copia = os.path.join(BACKUPS, "usuarios_%s.json" % marca)
                shutil.copy2(USUARIOS, copia)
                self._podar()
            tmp = USUARIOS + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(self.datos, f, ensure_ascii=False, indent=2)
            os.replace(tmp, USUARIOS)
        except OSError as e:
            messagebox.showerror("No se pudo guardar", str(e))

    @staticmethod
    def _podar():
        """Deja solo las ultimas MAX_BACKUPS copias."""
        try:
            copias = sorted(os.path.join(BACKUPS, f) for f in os.listdir(BACKUPS)
                            if f.startswith("usuarios_") and f.endswith(".json"))
            for viejo in copias[:-MAX_BACKUPS]:
                os.remove(viejo)
        except OSError:
            pass

    # -- tabla
    def refrescar(self):
        self.tabla.delete(*self.tabla.get_children())
        for i, u in enumerate(self.datos):
            clave = u.get("clave", "")
            if clave and not self.ver_claves.get():
                clave = "*" * len(clave)
            est = (u.get("status") or "").lower()
            tag = "mal" if (u.get("falla") or est in ("bloqueado", "quemado",
                                                      "con falla")) else (
                  "libre" if est == "disponible" else "")
            self.tabla.insert("", "end", iid=str(i), tags=(tag,),
                              values=(u.get("tipo", ""), u.get("usuario", ""), clave,
                                      u.get("descripcion", ""), u.get("status", ""),
                                      u.get("falla", "")))
        self.estado.set("%d usuario(s)  |  %s  |  copias en backups/"
                        % (len(self.datos), USUARIOS))

    def _sel(self):
        sel = self.tabla.selection()
        return int(sel[0]) if sel else None

    def cargar_en_form(self):
        i = self._sel()
        if i is None:
            return
        self.editando = i
        for c in self.COLS:
            self.vars[c].set(self.datos[i].get(c, ""))
        self.b_guardar.configure(text="Actualizar")

    def limpiar(self):
        self.editando = None
        for v in self.vars.values():
            v.set("")
        self.vars["tipo"].set(self.TIPOS[0])
        self.tabla.selection_remove(*self.tabla.selection())
        self.b_guardar.configure(text="Agregar")

    # -- acciones
    def guardar(self):
        fila = {c: self.vars[c].get().strip() for c in self.COLS}
        if not fila["usuario"]:
            messagebox.showinfo("Falta un dato", "Escribe el usuario.")
            return
        if self.editando is None:
            self.datos.append(fila)
        else:
            self.datos[self.editando] = fila
        self.volcar()
        self.limpiar()
        self.refrescar()

    def eliminar(self):
        i = self._sel()
        if i is None:
            messagebox.showinfo("Sin seleccion", "Escoge un usuario de la lista.")
            return
        if not messagebox.askyesno("Eliminar",
                                   "Quitar a %s de la lista?" % self.datos[i].get("usuario", "")):
            return
        self.datos.pop(i)
        self.volcar()
        self.limpiar()
        self.refrescar()

    def traer(self):
        """Agrega los del catalogo del repo que no tengas todavia."""
        compartidos = _json_o_nada(USUARIOS_COMPARTIDOS)
        if not isinstance(compartidos, list):
            messagebox.showinfo("Sin catalogo",
                                "No hay %s en el repo."
                                % os.path.basename(USUARIOS_COMPARTIDOS))
            return
        tengo = {self._clave(u) for u in self.datos}
        nuevos = [u for u in compartidos if self._clave(u) not in tengo]
        if not nuevos:
            self.estado.set("El catalogo del repo no trae ninguno nuevo.")
            return
        self.datos.extend(nuevos)
        self.volcar()
        self.refrescar()
        self.estado.set("%d usuario(s) traidos del repo." % len(nuevos))

    def publicar(self):
        """Escribe el catalogo que se versiona. La clave va solo si lo pediste."""
        con_claves = self.publicar_claves.get()
        aviso = ("Se van a publicar %d usuario(s) en %s, que SI se versiona.\n\n"
                 % (len(self.datos), os.path.basename(USUARIOS_COMPARTIDOS)))
        aviso += ("Las contrasenas van incluidas y quedaran en el historial de "
                  "git de forma permanente, aunque despues borres el archivo."
                  if con_claves else
                  "Las contrasenas NO van: cada quien completa las suyas.")
        if not messagebox.askyesno("Publicar al repo", aviso + "\n\nSeguir?"):
            return
        salida = []
        for u in self.datos:
            fila = {c: u.get(c, "") for c in self.COLS}
            if not con_claves:
                fila["clave"] = ""
            salida.append(fila)
        try:
            with open(USUARIOS_COMPARTIDOS, "w", encoding="utf-8") as f:
                json.dump(salida, f, ensure_ascii=False, indent=2)
        except OSError as e:
            messagebox.showerror("No se pudo publicar", str(e))
            return
        self.estado.set("%d usuario(s) publicados%s. Falta commitear el archivo."
                        % (len(salida), " con contrasena" if con_claves else
                           " sin contrasenas"))

    def abrir_datos(self):
        os.makedirs(BACKUPS, exist_ok=True)
        if hasattr(os, "startfile"):
            os.startfile(DIR_DATOS)      # Windows
        else:
            webbrowser.open("file:///" + DIR_DATOS.replace("\\", "/"))

    def copiar(self, campo):
        i = self._sel()
        valor = self.datos[i].get(campo, "") if i is not None else self.vars[campo].get()
        if not valor:
            return
        self.clipboard_clear()
        self.clipboard_append(valor)
        self.estado.set("Copiado al portapapeles.")


# --- ventana ---------------------------------------------------------------
class Panel:
    def __init__(self, raiz):
        self.raiz = raiz
        raiz.title("Panel QA - Colsubsidio")
        raiz.geometry("960x640")
        raiz.minsize(780, 480)

        self.tabs = ttk.Notebook(raiz)
        self.tabs.pack(fill="both", expand=True)

        self.herramientas = {}
        for spec in HERRAMIENTAS:
            raiz_h = spec.get("cwd", AQUI)
            objetivo = raiz_h if spec.get("comando") else os.path.join(AQUI, spec["script"])
            if not os.path.exists(objetivo):
                continue          # declarada pero sin script/repo: se omite
            h = Herramienta(self.tabs, spec)
            self.tabs.add(h, text="  %s  " % spec["nombre"])
            self.herramientas[spec["id"]] = h

        self.validaciones = None
        if os.path.isdir(FRAMEWORK):
            self.validaciones = Validaciones(self.tabs)
            self.tabs.add(self.validaciones, text="  Validaciones API  ")

        self.usuarios = Usuarios(self.tabs)
        self.tabs.add(self.usuarios, text="  Usuarios  ")

        self.corridas = None
        if "observador" in self.herramientas:
            self.corridas = Corridas(self.tabs, self)
            self.tabs.add(self.corridas, text="  Corridas  ")
            self.herramientas["observador"].al_terminar = [self.corridas.refrescar]

        raiz.protocol("WM_DELETE_WINDOW", self.cerrar)
        raiz.after(120, self.bombear)

    def bombear(self):
        """Vuelca a cada log lo que su hilo fue dejando en la cola. Los widgets
        de tkinter solo se tocan desde el hilo principal."""
        for h in self.herramientas.values():
            h.drenar()
        if self.validaciones:
            self.validaciones.drenar()
        self.raiz.after(120, self.bombear)

    def cerrar(self):
        vivas = [h for h in self.herramientas.values() if h.ocupada()]
        if self.validaciones and self.validaciones.ocupada():
            vivas.append(self.validaciones)
        if vivas:
            nombres = ", ".join(h.spec["nombre"] for h in vivas)
            if not messagebox.askyesno(
                    "Procesos en curso",
                    "Sigue corriendo: %s.\n\nSi cierras ahora se cortan "
                    "(una captura quedaria sin reporte).\n\nCerrar de todas "
                    "formas?" % nombres):
                return
            for h in vivas:
                h.proc.terminate()
        self.raiz.destroy()


def main():
    raiz = tk.Tk()
    try:
        ttk.Style().theme_use("vista")
    except tk.TclError:
        pass
    Panel(raiz)
    raiz.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())
